import io
from typing import List, Dict, Any, Optional
from app.modules.data_output.exporters.base_exporter import BaseExporter

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


class ExcelExporter(BaseExporter):
    """Excel/XLSX document exporter using OpenPyXL."""

    async def export_query_results(
        self,
        columns: List[str],
        rows: List[List[Any]],
        title: str = "Query Results",
        options: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        if openpyxl is not None:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = title[:31].replace("/", "_").replace("\\", "_")

            # Header styling (Navy blue fill, bold white font)
            header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            header_align = Alignment(horizontal="center", vertical="center")

            # Write header row
            ws.append(columns)
            for col_num in range(1, len(columns) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align

            # Write data rows
            for row in rows:
                formatted_row = [str(item) if item is not None else "" for item in row]
                ws.append(formatted_row)

            # Auto-fit column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        # Simple XML fallback if openpyxl is unavailable
        xml_str = '<?xml version="1.0"?><?mso-application progid="Excel.Sheet"?>'
        xml_str += '<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"'
        xml_str += ' xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">\n<Worksheet ss:Name="Results"><Table>\n'
        xml_str += '<Row>' + ''.join(f'<Cell><Data ss:Type="String">{c}</Data></Cell>' for c in columns) + '</Row>\n'
        for row in rows:
            xml_str += '<Row>'
            for item in row:
                val = str(item) if item is not None else ""
                xml_str += f'<Cell><Data ss:Type="String">{val}</Data></Cell>'
            xml_str += '</Row>\n'
        xml_str += '</Table></Worksheet></Workbook>'
        return xml_str.encode("utf-8")

    async def export_report(
        self,
        title: str,
        content_structure: Dict[str, Any],
        options: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        sections = content_structure.get("sections", [])
        sec_rows = []
        for s in sections:
            sec_rows.append([s.get("title", "Section"), s.get("content", "")])

        return await self.export_query_results(
            columns=["Section Title", "Content"],
            rows=sec_rows,
            title=title,
            options=options,
        )
